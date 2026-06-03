"""
admin_api/cargo_views.py
------------------------
Cargo bo'limi uchun to'liq API:
- Yakka yuk qo'shish
- Excel import (TREK RAQAM, REYS, ID, OMBORDA)
- Bulk status o'zgartirish (Yo'lga, Punktga, Topshirildi)
- Push notification
- Tahrirlash, o'chirish
"""
import re
import openpyxl
from io import BytesIO
from django.utils import timezone
from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from cargo.models import Cargo
from accounts.models import User


# ─── YORDAMCHI FUNKSIYALAR ───────────────────────────────────

def get_transport_type(id_value):
    if not id_value:
        return None
    clean = re.sub(r'[\s\-]', '', str(id_value).strip().upper())
    if clean.startswith('US') and re.match(r'^US\d', clean):
        return 'AVIA'
    elif clean.startswith('GG') and re.match(r'^GG\d', clean):
        return 'AVTO'
    return None


def extract_number(id_value):
    if not id_value:
        return None
    match = re.search(r'(\d+)', str(id_value).strip())
    return str(int(match.group(1))) if match else None


def find_user_by_number(number, cache=None):
    if not number:
        return None
    if cache is not None and number in cache:
        return cache[number]

    num = str(int(number))
    num_padded = num.zfill(4)
    patterns = []
    for n in [num, num_padded]:
        patterns += [f"US-{n}", f"US{n}", f"US {n}", f"GG-{n}", f"GG{n}", f"GG {n}"]
    patterns = list(dict.fromkeys(patterns))

    user = None
    for p in patterns:
        user = User.objects.filter(user_id__iexact=p).first()
        if user:
            break
    if not user:
        user = User.objects.filter(user_id__icontains=num).first()

    if cache is not None:
        cache[number] = user
    return user


def send_push(cargos, status):
    """Push notification yuborish"""
    try:
        from utils.push_service import send_flight_status_push
        success, error = send_flight_status_push(cargos, status)
        return success, error
    except Exception as e:
        print(f"Push xatosi: {e}")
        return 0, 0


def cargo_dict(c):
    return {
        "id": c.id,
        "track_code": c.track_code,
        "flight_name": c.flight_name or "-",
        "status": c.status,
        "transport_type": c.transport_type or "-",
        "created_at": c.created_at.strftime("%d.%m.%Y"),
        "delivered_at": c.delivered_at.strftime("%d.%m.%Y %H:%M") if c.delivered_at else None,
        "user_id": c.user.user_id if c.user else "-",
        "user_name": f"{c.user.first_name} {c.user.last_name}".strip() if c.user else "Topilmadi",
        "user_phone": c.user.phone if c.user else "-",
        "warehouse_admin": str(c.warehouse_admin) if c.warehouse_admin else None,
        "onway_admin": str(c.onway_admin) if c.onway_admin else None,
        "arrived_admin": str(c.arrived_admin) if c.arrived_admin else None,
        "delivered_admin": str(c.delivered_admin) if c.delivered_admin else None,
    }


# ─── RO'YXAT ─────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def cargo_list(request):
    """GET /api/admin/cargos/"""
    qs = Cargo.objects.select_related("user").order_by("-created_at")

    if s := request.GET.get("status", ""):
        qs = qs.filter(status=s)
    if t := request.GET.get("transport", ""):
        qs = qs.filter(transport_type=t)
    if f := request.GET.get("flight", ""):
        qs = qs.filter(flight_name__icontains=f)
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(
            Q(track_code__icontains=q) |
            Q(user__user_id__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__phone__icontains=q) |
            Q(flight_name__icontains=q)
        )

    total = qs.count()
    page = int(request.GET.get("page", 1))
    size = 50
    results = [cargo_dict(c) for c in qs[(page - 1) * size: page * size]]
    return Response({"results": results, "total": total, "page": page, "pages": (total + size - 1) // size})


# ─── YAKKA QO'SHISH ──────────────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAdminUser])
def cargo_create(request):
    """
    POST /api/admin/cargos/create/
    Body: { track_code, flight_name, user_id, transport_type, created_at }
    """
    track_code = request.data.get("track_code", "").strip()
    if not track_code:
        return Response({"error": "Trek kodi kiritilmagan"}, status=400)

    if Cargo.objects.filter(track_code=track_code).exists():
        return Response({"error": f"Bu trek kodi allaqachon mavjud: {track_code}"}, status=400)

    # Foydalanuvchini topish
    user_id_str = request.data.get("user_id", "").strip()
    user = None
    if user_id_str:
        user = User.objects.filter(user_id__iexact=user_id_str).first()
        if not user:
            number = extract_number(user_id_str)
            user = find_user_by_number(number)

    transport = request.data.get("transport_type") or get_transport_type(user_id_str)
    status = "Omborda" if user else "Kutilmoqda"

    cargo = Cargo(
        track_code=track_code,
        flight_name=request.data.get("flight_name", "").strip() or None,
        user=user,
        status=status,
        transport_type=transport,
        created_by=request.user,
        updated_by=request.user,
    )
    if status == "Omborda":
        cargo.warehouse_admin = request.user

    if created_at := request.data.get("created_at"):
        try:
            from django.utils.dateparse import parse_date
            d = parse_date(created_at)
            if d:
                cargo.created_at = timezone.make_aware(timezone.datetime.combine(d, timezone.datetime.min.time()))
        except Exception:
            pass

    cargo._skip_push_signal = False
    cargo.save()

    return Response({"ok": True, "id": cargo.id, "status": status, "user_found": user is not None})


# ─── EXCEL IMPORT ────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAdminUser])
@parser_classes([MultiPartParser, FormParser])
def cargo_import_excel(request):
    """
    POST /api/admin/cargos/import/
    File: excel (xlsx)
    Ustunlar: TREK RAQAM | REYS | ID | OMBORDA
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "Fayl yuklanmagan"}, status=400)

    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return Response({"error": f"Faylni o'qib bo'lmadi: {e}"}, status=400)

    # Ustun sarlavhalarini topish
    headers = {}
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_vals = [str(v).strip().upper() if v else "" for v in row]
        for j, val in enumerate(row_vals):
            if "TREK" in val or "ТРЕК" in val:
                headers["track_code"] = j
            if "REYS" in val or "РЕЙС" in val:
                headers["flight_name"] = j
            if val == "ID":
                headers["id"] = j
            if "OMBOR" in val or "ОМБОР" in val or "SANA" in val:
                headers["created_at"] = j
        if "track_code" in headers:
            header_row = i
            break

    if "track_code" not in headers:
        return Response({"error": "TREK RAQAM ustuni topilmadi. Ustun nomlari: TREK RAQAM, REYS, ID, OMBORDA"}, status=400)

    user_cache = {}
    results = {"imported": 0, "pending": 0, "skipped": 0, "errors": [], "pending_list": []}

    # Reys nomini aniqlash (birinchi qatordan)
    default_flight = request.data.get("flight_name", "").strip() or None

    new_cargos = []

    for row in ws.iter_rows(min_row=(header_row or 1) + 1, values_only=True):
        track_code = row[headers["track_code"]] if headers.get("track_code") is not None and len(row) > headers["track_code"] else None
        if not track_code or str(track_code).strip().lower() in ["", "none", "null"]:
            continue

        track_code = str(track_code).strip()

        # Mavjudligini tekshirish
        if Cargo.objects.filter(track_code=track_code).exists():
            results["skipped"] += 1
            continue

        # Reys
        flight_name = None
        if headers.get("flight_name") is not None and len(row) > headers["flight_name"]:
            flight_name = str(row[headers["flight_name"]]).strip() if row[headers["flight_name"]] else None
        flight_name = flight_name or default_flight or "Noma'lum"

        # ID va foydalanuvchi
        id_value = None
        if headers.get("id") is not None and len(row) > headers["id"]:
            id_value = str(row[headers["id"]]).strip() if row[headers["id"]] else None

        number = extract_number(id_value)
        transport = get_transport_type(id_value)
        user = find_user_by_number(number, user_cache) if number else None

        # Sana
        created_at = timezone.now()
        if headers.get("created_at") is not None and len(row) > headers["created_at"]:
            val = row[headers["created_at"]]
            if val:
                try:
                    if hasattr(val, 'date'):
                        created_at = timezone.make_aware(timezone.datetime.combine(val.date() if hasattr(val, 'date') else val, timezone.datetime.min.time()))
                except Exception:
                    pass

        cargo = Cargo(
            track_code=track_code,
            flight_name=flight_name,
            user=user,
            status="Omborda" if user else "Kutilmoqda",
            transport_type=transport,
            created_at=created_at,
            created_by=request.user,
            updated_by=request.user,
        )
        if user:
            cargo.warehouse_admin = request.user
        cargo._skip_push_signal = True

        try:
            cargo.save()
            new_cargos.append(cargo)
            if user:
                results["imported"] += 1
            else:
                results["pending"] += 1
                results["pending_list"].append({"track_code": track_code, "id_value": id_value or "-"})
        except Exception as e:
            results["errors"].append({"track_code": track_code, "error": str(e)})

    # Push yuborish
    cargos_with_user = [c for c in new_cargos if c.user]
    if cargos_with_user:
        success, error = send_push(cargos_with_user, "Omborda")
        results["push_sent"] = success
        results["push_error"] = error

    return Response({
        "ok": True,
        "imported": results["imported"],
        "pending": results["pending"],
        "skipped": results["skipped"],
        "push_sent": results.get("push_sent", 0),
        "errors": results["errors"][:10],
        "pending_list": results["pending_list"][:20],
    })


# ─── YAKKA TAHRIRLASH ────────────────────────────────────────

@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAdminUser])
def cargo_detail(request, pk):
    """GET/PATCH/DELETE /api/admin/cargos/<pk>/"""
    try:
        cargo = Cargo.objects.select_related("user").get(pk=pk)
    except Cargo.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)

    if request.method == "GET":
        return Response(cargo_dict(cargo))

    if request.method == "DELETE":
        cargo.delete()
        return Response({"ok": True})

    # PATCH
    if "track_code" in request.data:
        new_tc = request.data["track_code"].strip()
        if Cargo.objects.exclude(pk=pk).filter(track_code=new_tc).exists():
            return Response({"error": "Bu trek kodi allaqachon mavjud"}, status=400)
        cargo.track_code = new_tc

    if "flight_name" in request.data:
        cargo.flight_name = request.data["flight_name"].strip() or None

    if "transport_type" in request.data:
        cargo.transport_type = request.data["transport_type"] or None

    # Foydalanuvchi o'zgartirish
    if "user_id" in request.data:
        uid = request.data["user_id"].strip()
        if uid:
            u = User.objects.filter(user_id__iexact=uid).first()
            if not u:
                return Response({"error": f"Foydalanuvchi topilmadi: {uid}"}, status=400)
            cargo.user = u
        else:
            cargo.user = None

    cargo.updated_by = request.user
    cargo._skip_push_signal = True
    cargo.save()
    return Response({"ok": True, "cargo": cargo_dict(cargo)})


# ─── STATUS O'ZGARTIRISH (YAKKA) ─────────────────────────────

@api_view(["PATCH"])
@permission_classes([IsAdminUser])
def cargo_update_status(request, pk):
    """PATCH /api/admin/cargos/<pk>/status/"""
    try:
        cargo = Cargo.objects.get(pk=pk)
    except Cargo.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)

    new_status = request.data.get("status")
    valid = ["Kutilmoqda", "Omborda", "Yo'lda", "Punktda", "Topshirildi"]
    if new_status not in valid:
        return Response({"error": "Noto'g'ri status"}, status=400)

    cargo.status = new_status
    cargo.updated_by = request.user
    if new_status == "Omborda":
        cargo.warehouse_admin = request.user
    elif new_status == "Yo'lda":
        cargo.onway_admin = request.user
    elif new_status == "Punktda":
        cargo.arrived_admin = request.user
    elif new_status == "Topshirildi":
        cargo.delivered_admin = request.user
        cargo.delivered_at = timezone.now()

    cargo._skip_push_signal = False  # Push yuborilsin
    cargo.save()
    return Response({"ok": True, "status": cargo.status})


# ─── BULK STATUS O'ZGARTIRISH ────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAdminUser])
def cargo_bulk_status(request):
    """
    POST /api/admin/cargos/bulk-status/
    Body: { ids: [1,2,3], status: "Yo'lda" }
    """
    ids = request.data.get("ids", [])
    new_status = request.data.get("status")

    valid = ["Kutilmoqda", "Omborda", "Yo'lda", "Punktda", "Topshirildi"]
    if new_status not in valid:
        return Response({"error": "Noto'g'ri status"}, status=400)
    if not ids:
        return Response({"error": "Yuklar tanlanmagan"}, status=400)

    qs = Cargo.objects.filter(pk__in=ids)
    count = qs.count()

    update_fields = {"status": new_status, "updated_by": request.user}
    if new_status == "Omborda":
        update_fields["warehouse_admin"] = request.user
    elif new_status == "Yo'lda":
        update_fields["onway_admin"] = request.user
    elif new_status == "Punktda":
        update_fields["arrived_admin"] = request.user
    elif new_status == "Topshirildi":
        update_fields["delivered_admin"] = request.user
        update_fields["delivered_at"] = timezone.now()

    qs.update(**update_fields)

    # Push
    updated = list(Cargo.objects.filter(pk__in=ids).select_related("user"))
    push_ok, push_err = send_push(updated, new_status)

    return Response({
        "ok": True,
        "updated": count,
        "push_sent": push_ok,
        "push_error": push_err,
    })


# ─── SELECT ALL IDS (BULK UCHUN) ─────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def cargo_all_ids(request):
    """
    GET /api/admin/cargos/all-ids/?status=Omborda
    Joriy filter bo'yicha barcha ID lar
    """
    qs = Cargo.objects.all()
    if s := request.GET.get("status"):
        qs = qs.filter(status=s)
    if t := request.GET.get("transport"):
        qs = qs.filter(transport_type=t)
    if f := request.GET.get("flight"):
        qs = qs.filter(flight_name__icontains=f)
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(
            Q(track_code__icontains=q) | Q(user__user_id__icontains=q) | Q(flight_name__icontains=q)
        )
    ids = list(qs.values_list("id", flat=True))
    return Response({"ids": ids, "total": len(ids)})


# ─── EXCEL EXPORT ────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def cargo_export_excel(request):
    """GET /api/admin/cargos/export/?status=Omborda"""
    from django.http import HttpResponse

    qs = Cargo.objects.select_related("user").order_by("-created_at")
    if s := request.GET.get("status"):
        qs = qs.filter(status=s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yuklar"
    ws.append(["TREK RAQAM", "REYS", "ID", "OMBORDA", "STATUS", "TRANSPORT", "FOYDALANUVCHI"])

    for c in qs:
        ws.append([
            c.track_code,
            c.flight_name or "",
            c.user.user_id if c.user else "",
            c.created_at.strftime("%d.%m.%Y"),
            c.status,
            c.transport_type or "",
            f"{c.user.first_name} {c.user.last_name}".strip() if c.user else "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="yuklar.xlsx"'
    return response
