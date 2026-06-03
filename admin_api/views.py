"""
admin_api/views.py  —  To'liq React Admin Panel API
"""
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.contrib.auth import authenticate
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAdminUser, AllowAny
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework_simplejwt.tokens import RefreshToken
from datetime import timedelta

from cargo.models import Cargo
from warehouse.models import ArrivedGroup
from accounts.models import User, OTPCode
from flights.models import Flight
from services.models import SupportMessage, CalculationRequest, WarehouseSettings, AppVersion, TutorialVideo
from notifications.models import Notification
from unassigned.models import UnassignedCargo


# ── AUTH ─────────────────────────────────────

@api_view(["POST"])
@permission_classes([AllowAny])
def admin_login(request):
    phone = request.data.get("phone", "").strip()
    password = request.data.get("password", "")
    try:
        u = User.objects.get(phone=phone)
    except User.DoesNotExist:
        return Response({"error": "Foydalanuvchi topilmadi"}, status=401)
    if not u.check_password(password):
        return Response({"error": "Parol xato"}, status=401)
    if not u.is_staff:
        return Response({"error": "Siz admin emassiz"}, status=403)
    refresh = RefreshToken.for_user(u)
    return Response({
        "access": str(refresh.access_token),
        "refresh": str(refresh),
        "user": {"id": u.id, "name": f"{u.first_name} {u.last_name}".strip() or u.phone, "phone": u.phone, "is_superuser": u.is_superuser},
    })


# ── DASHBOARD ────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def dashboard_stats(request):
    """
    Dashboard uchun to'liq statistika.
    Barcha hisob-kitob delivered=Topshirildi guruhlardan olinadi.
    Oylik grafik: so'mda (kassa naqd + kassa karta + ilovadan tasdiqlangan)
    """
    from services.models import WarehouseSettings
    import datetime

    now = timezone.now()

    # Dollar kursi
    try:
        ws = WarehouseSettings.objects.filter(pk=1).first()
        rate = float(getattr(ws, "dollar_rate", None) or 12700)
    except Exception:
        rate = 12700.0

    # ── TOPSHIRILGAN GURUHLAR ──────────────────────────────────
    delivered = ArrivedGroup.objects.filter(payment_status="Topshirildi")

    # Kassadan to'lovlar (cash_amount yoki card_amount bor)
    kassa_cash_total = 0.0
    kassa_card_total = 0.0
    confirmed_usd_total = 0.0

    for g in delivered.values("total_price", "cash_amount", "card_amount"):
        cash = float(g["cash_amount"] or 0)
        card = float(g["card_amount"] or 0)
        usd  = float(g["total_price"] or 0)
        if cash > 0 or card > 0:
            kassa_cash_total += cash
            kassa_card_total += card
        else:
            confirmed_usd_total += usd

    confirmed_sum_total = round(confirmed_usd_total * rate)
    total_som = round(kassa_cash_total + kassa_card_total + confirmed_sum_total)
    total_cash = round(kassa_cash_total)
    total_card = round(kassa_card_total + confirmed_sum_total)  # ilovadan to'langan = "karta" sifatida

    # ── YUK STATISTIKASI ──────────────────────────────────────
    cargo_stats = Cargo.objects.aggregate(
        total=Count("id"),
        warehouse=Count("id", filter=Q(status="Omborda")),
        onway=Count("id", filter=Q(status="Yo'lda")),
        arrived=Count("id", filter=Q(status="Punktda")),
        delivered=Count("id", filter=Q(status="Topshirildi")),
        pending=Count("id", filter=Q(status="Kutilmoqda")),
    )

    # ── FOYDALANUVCHI STATISTIKASI ────────────────────────────
    user_stats = User.objects.filter(is_staff=False).aggregate(
        total=Count("id"),
        approved=Count("id", filter=Q(status="approved")),
        pending=Count("id", filter=Q(status="pending")),
        rejected=Count("id", filter=Q(status="rejected")),
    )

    # ── OGOHLANTIRISHLAR ──────────────────────────────────────
    pending_payments = ArrivedGroup.objects.filter(payment_status="Tasdiqlash jarayonida").count()
    pending_calc = CalculationRequest.objects.filter(is_responded=False).count()

    # ── OYLIK GRAFIK (oxirgi 6 oy, so'mda) ───────────────────
    month_names = ["Yan", "Feb", "Mar", "Apr", "May", "Iyn", "Iyl", "Avg", "Sen", "Okt", "Noy", "Dek"]
    monthly_data = []

    for i in range(5, -1, -1):
        # Oyning birinchi kunini aniqlaymiz
        ms = (now.replace(day=1) - datetime.timedelta(days=i * 30)).replace(day=1)
        me = (ms + datetime.timedelta(days=32)).replace(day=1)

        # Shu oyda topshirilgan guruhlar
        # delivered_at bo'lsa uni, yo'q bo'lsa created_at ni ishlatamiz
        month_groups = delivered.filter(
            Q(delivered_at__gte=ms, delivered_at__lt=me) |
            Q(delivered_at__isnull=True, created_at__gte=ms, created_at__lt=me)
        )

        m_cash = 0.0
        m_card = 0.0
        for g in month_groups.values("total_price", "cash_amount", "card_amount"):
            cash = float(g["cash_amount"] or 0)
            card = float(g["card_amount"] or 0)
            usd  = float(g["total_price"] or 0)
            if cash > 0 or card > 0:
                m_cash += cash
                m_card += card
            else:
                # ilovadan to'langan USD → so'mga
                m_card += round(usd * rate)

        monthly_data.append({
            "month": month_names[ms.month - 1],
            "naqd": round(m_cash),
            "karta": round(m_card),
            "jami": round(m_cash + m_card),
        })

    # ── SO'NGGI TO'LOVLAR ────────────────────────────────────
    recent_payments = []
    for g in ArrivedGroup.objects.select_related("user", "delivered_admin", "created_by").order_by("-created_at")[:10]:
        cash = float(g.cash_amount or 0) if hasattr(g, "cash_amount") else 0
        card = float(g.card_amount or 0) if hasattr(g, "card_amount") else 0
        if cash > 0 and card == 0:
            method = "naqd"
        elif card > 0 and cash == 0:
            method = "karta"
        elif cash > 0 and card > 0:
            method = "aralash"
        else:
            method = "karta"  # ilovadan to'langan

        # Yetkazish usuli ko'rsatilsin
        delivery = g.delivery_method or "Punktda"
        delivery_labels = {
            "Punktda": "🏪 Punktdan olib ketdi",
            "Pochta": "📮 Pochta orqali",
            "Taksi": "🚕 Taksi orqali",
        }

        recent_payments.append({
            "id": g.id,
            "receipt_code": g.receipt_code,
            "user_id": g.user.user_id or "-",
            "user_name": f"{g.user.first_name} {g.user.last_name}".strip() or g.user.phone,
            "amount": float(g.total_price),
            "amount_sum": round(float(g.total_price) * rate),
            "method": method,
            "delivery_method": delivery,
            "delivery_label": delivery_labels.get(delivery, delivery),
            "status": g.payment_status,
            "admin_name": str(g.delivered_admin) if g.delivered_admin else "-",
            "created_admin": str(g.created_by) if g.created_by else "-",
            "date": g.created_at.strftime("%d.%m.%Y"),
            "delivered_at": g.delivered_at.strftime("%d.%m.%Y %H:%M") if g.delivered_at else None,
        })

    return Response({
        "income": {
            "total": total_som,
            "card": total_card,
            "cash": total_cash,
            "confirmed_usd": confirmed_usd_total,
            "confirmed_sum": confirmed_sum_total,
            "kassa_cash": round(kassa_cash_total),
            "kassa_card": round(kassa_card_total),
        },
        "cargos": cargo_stats,
        "users": user_stats,
        "alerts": {
            "pending_payments": pending_payments,
            "pending_users": user_stats["pending"],
            "pending_calc": pending_calc,
        },
        "monthly_data": monthly_data,
        "recent_payments": recent_payments,
        "dollar_rate": rate,
    })


# ── YUKLAR ───────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def cargo_list(request):
    qs = Cargo.objects.select_related("user").order_by("-created_at")
    if s := request.GET.get("status"): qs = qs.filter(status=s)
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(Q(track_code__icontains=q) | Q(user__user_id__icontains=q) | Q(flight_name__icontains=q))
    if t := request.GET.get("transport"): qs = qs.filter(transport_type=t)

    total = qs.count()
    page = int(request.GET.get("page", 1))
    size = 30
    results = []
    for c in qs[(page - 1) * size: page * size]:
        results.append({
            "id": c.id, "track_code": c.track_code, "flight_name": c.flight_name or "-",
            "status": c.status, "transport_type": c.transport_type or "-",
            "created_at": c.created_at.strftime("%d.%m.%Y"),
            "user_id": c.user.user_id if c.user else "-",
            "user_name": f"{c.user.first_name} {c.user.last_name}".strip() if c.user else "Topilmadi",
            "user_phone": c.user.phone if c.user else "-",
        })
    return Response({"results": results, "total": total, "page": page, "pages": (total + size - 1) // size})


@api_view(["PATCH"])
@permission_classes([IsAdminUser])
def cargo_update_status(request, pk):
    try:
        cargo = Cargo.objects.get(pk=pk)
    except Cargo.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    new_status = request.data.get("status")
    if new_status not in ["Kutilmoqda", "Omborda", "Yo'lda", "Punktda", "Topshirildi"]:
        return Response({"error": "Noto'g'ri status"}, status=400)
    cargo.status = new_status
    if new_status == "Omborda": cargo.warehouse_admin = request.user
    elif new_status == "Yo'lda": cargo.onway_admin = request.user
    elif new_status == "Punktda": cargo.arrived_admin = request.user
    elif new_status == "Topshirildi": cargo.delivered_admin = request.user; cargo.delivered_at = timezone.now()
    cargo.updated_by = request.user
    cargo.save()
    return Response({"ok": True})


# ── FOYDALANUVCHILAR ─────────────────────────

def _user_data(u):
    return {
        "id": u.id, "user_id": u.user_id or "-", "phone": u.phone,
        "full_name": f"{u.first_name} {u.last_name}".strip() or "-",
        "first_name": u.first_name, "last_name": u.last_name,
        "status": u.status, "is_active": u.is_active,
        "date_joined": u.date_joined.strftime("%d.%m.%Y"),
        "last_active": u.last_active.strftime("%d.%m.%Y %H:%M") if u.last_active else "-",
        "passport_series": u.passport_series or "-",
        "cargo_count": u.cargos.count(),
    }


@api_view(["GET"])
@permission_classes([IsAdminUser])
def user_list(request):
    qs = User.objects.filter(is_staff=False, status="approved").order_by("-date_joined")
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(Q(phone__icontains=q) | Q(user_id__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))
    total = qs.count()
    page = int(request.GET.get("page", 1))
    size = 30
    return Response({"results": [_user_data(u) for u in qs[(page - 1) * size: page * size]], "total": total, "page": page, "pages": (total + size - 1) // size})


@api_view(["GET"])
@permission_classes([IsAdminUser])
def application_list(request):
    """Kutilayotgan va rad etilgan arizalar"""
    qs = User.objects.filter(is_staff=False).exclude(status="approved").order_by("-date_joined")
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(Q(phone__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))
    total = qs.count()
    page = int(request.GET.get("page", 1))
    size = 30
    return Response({"results": [_user_data(u) for u in qs[(page - 1) * size: page * size]], "total": total, "page": page, "pages": (total + size - 1) // size})


@api_view(["GET"])
@permission_classes([IsAdminUser])
def user_detail(request, pk):
    try:
        u = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    cargos = Cargo.objects.filter(user=u).order_by("-created_at")[:10]
    data = _user_data(u)
    data.update({
        "jshshir": u.jshshir or "-", "passport_series": u.passport_series or "-",
        "birth_date": str(u.birth_date) if u.birth_date else "-",
        "address": u.address or "-",
        "rejection_reason": u.rejection_reason or "",
        "rejection_note": u.rejection_note or "",
        "passport_front": request.build_absolute_uri(u.passport_front.url) if u.passport_front else None,
        "passport_back": request.build_absolute_uri(u.passport_back.url) if u.passport_back else None,
        "recent_cargos": [{"track_code": c.track_code, "status": c.status, "flight_name": c.flight_name, "created_at": c.created_at.strftime("%d.%m.%Y")} for c in cargos],
    })
    return Response(data)


@api_view(["PATCH"])
@permission_classes([IsAdminUser])
def user_approve(request, pk):
    try:
        u = User.objects.get(pk=pk, is_staff=False)
    except User.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    if not u.user_id:
        max_num = 100
        for eu in User.objects.filter(user_id__startswith="UTS-"):
            try:
                n = int(eu.user_id.split("-")[1])
                if n > max_num: max_num = n
            except: pass
        u.user_id = f"UTS-{max_num + 1:04d}"
    u.status = "approved"; u.is_active = True; u.rejection_reason = None; u.rejection_note = None
    u.save()
    return Response({"ok": True, "user_id": u.user_id})


@api_view(["PATCH"])
@permission_classes([IsAdminUser])
def user_reject(request, pk):
    try:
        u = User.objects.get(pk=pk, is_staff=False)
    except User.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    u.status = "rejected"; u.rejection_reason = request.data.get("reason", "other"); u.rejection_note = request.data.get("note", "")
    u.save()
    return Response({"ok": True})


# ── OTP ──────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def otp_list(request):
    qs = OTPCode.objects.select_related("user").order_by("-created_at")
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(Q(user__phone__icontains=q) | Q(user__first_name__icontains=q))
    total = qs.count()
    page = int(request.GET.get("page", 1))
    size = 30
    results = []
    for o in qs[(page - 1) * size: page * size]:
        results.append({
            "id": o.id, "code": o.code,
            "phone": o.user.phone,
            "user_name": f"{o.user.first_name} {o.user.last_name}".strip() or o.user.phone,
            "created_at": o.created_at.strftime("%d.%m.%Y %H:%M"),
        })
    return Response({"results": results, "total": total, "page": page, "pages": (total + size - 1) // size})


# ── TO'LOVLAR ────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def payment_list(request):
    qs = ArrivedGroup.objects.select_related("user").order_by("-created_at")
    if s := request.GET.get("status"): qs = qs.filter(payment_status=s)
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(Q(receipt_code__icontains=q) | Q(user__user_id__icontains=q) | Q(user__phone__icontains=q))
    total = qs.count()
    page = int(request.GET.get("page", 1))
    size = 30
    results = []
    for g in qs[(page - 1) * size: page * size]:
        results.append({
            "id": g.id, "receipt_code": g.receipt_code,
            "user_id": g.user.user_id or "-",
            "user_name": f"{g.user.first_name} {g.user.last_name}".strip() or g.user.phone,
            "user_phone": g.user.phone,
            "total_price": float(g.total_price), "total_weight": float(g.total_weight),
            "payment_status": g.payment_status,
            "delivery_method": g.delivery_method or "-",
            "delivery_address": g.delivery_address or "-",
            "cargo_count": g.selected_cargos.count(),
            "payment_check": request.build_absolute_uri(g.payment_check.url) if g.payment_check else None,
            "admin_note": g.admin_note or "",
            "created_at": g.created_at.strftime("%d.%m.%Y %H:%M"),
        })
    return Response({"results": results, "total": total, "page": page, "pages": (total + size - 1) // size})


@api_view(["PATCH"])
@permission_classes([IsAdminUser])
def payment_approve(request, pk):
    try:
        g = ArrivedGroup.objects.get(pk=pk)
    except ArrivedGroup.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    g.payment_status = "To'lov tasdiqlandi"
    g.admin_note = f"Tasdiqlandi ✅ — {request.user}"
    g.save()
    return Response({"ok": True})


@api_view(["PATCH"])
@permission_classes([IsAdminUser])
def payment_reject(request, pk):
    try:
        g = ArrivedGroup.objects.get(pk=pk)
    except ArrivedGroup.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    g.payment_status = "To'lov rad etildi"
    g.admin_note = request.data.get("note", "Rad etildi ❌")
    g.save()
    return Response({"ok": True})


@api_view(["PATCH"])
@permission_classes([IsAdminUser])
def payment_deliver(request, pk):
    try:
        g = ArrivedGroup.objects.get(pk=pk)
    except ArrivedGroup.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    g.payment_status = "Topshirildi"
    g.delivered_admin = request.user
    g.save()
    cargos = g.selected_cargos.all()
    cargos.update(status="Topshirildi", delivered_at=timezone.now(), delivered_admin=request.user)
    return Response({"ok": True})


# ── REYSLAR ──────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def flight_list(request):
    data = []
    for f in Flight.objects.order_by("-arrival_date"):
        data.append({
            "id": f.id, "name": f.name, "status": f.status,
            "status_display": f.get_status_display(),
            "warehouse_start": str(f.warehouse_start),
            "warehouse_end": str(f.warehouse_end),
            "arrival_date": str(f.arrival_date),
            "note": f.note or "",
            "cargo_count": Cargo.objects.filter(flight_name=f.name).count(),
        })
    return Response(data)


@api_view(["POST"])
@permission_classes([IsAdminUser])
def flight_create(request):
    try:
        f = Flight.objects.create(
            name=request.data["name"],
            warehouse_start=request.data["warehouse_start"],
            warehouse_end=request.data["warehouse_end"],
            arrival_date=request.data["arrival_date"],
            status=request.data.get("status", "jarayonda"),
            note=request.data.get("note", ""),
        )
        return Response({"ok": True, "id": f.id})
    except Exception as e:
        return Response({"error": str(e)}, status=400)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAdminUser])
def flight_detail(request, pk):
    try:
        f = Flight.objects.get(pk=pk)
    except Flight.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    if request.method == "DELETE":
        f.delete(); return Response({"ok": True})
    for field in ["name", "status", "warehouse_start", "warehouse_end", "arrival_date", "note"]:
        if field in request.data: setattr(f, field, request.data[field])
    f.save()
    return Response({"ok": True})


# ── CHAT ─────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def chat_users(request):
    qs = User.objects.filter(is_staff=False, chat_messages__isnull=False).distinct()
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(phone__icontains=q))
    data = []
    for u in qs:
        unread = SupportMessage.objects.filter(user=u, is_from_admin=False, is_read=False).count()
        last = SupportMessage.objects.filter(user=u).order_by("-created_at").first()
        data.append({
            "id": u.id, "user_id": u.user_id or "-",
            "full_name": f"{u.first_name} {u.last_name}".strip() or u.phone,
            "phone": u.phone, "unread_count": unread,
            "last_message": last.message[:40] if last and last.message else "",
        })
    data.sort(key=lambda x: x["unread_count"], reverse=True)
    return Response(data)


@api_view(["GET"])
@permission_classes([IsAdminUser])
def chat_messages(request, user_id):
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    SupportMessage.objects.filter(user=user, is_from_admin=False).update(is_read=True)
    msgs = SupportMessage.objects.filter(user=user).order_by("created_at")
    data = []
    for m in msgs:
        data.append({
            "id": m.id, "message": m.message or "",
            "image": request.build_absolute_uri(m.image.url) if m.image else None,
            "is_from_admin": m.is_from_admin,
            "created_at": m.created_at.strftime("%H:%M %d.%m"),
        })
    return Response(data)


@api_view(["POST"])
@permission_classes([IsAdminUser])
@parser_classes([MultiPartParser, FormParser])
def chat_reply(request, user_id):
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    msg = SupportMessage.objects.create(
        user=user, admin=request.user,
        message=request.data.get("message", "") or None,
        image=request.FILES.get("image"),
        is_from_admin=True,
    )
    return Response({"ok": True, "id": msg.id})


# ── KALKULATOR ───────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def calc_list(request):
    qs = CalculationRequest.objects.select_related("user").order_by("-created_at")
    if r := request.GET.get("responded"):
        qs = qs.filter(is_responded=(r == "true"))
    data = []
    for c in qs[:50]:
        data.append({
            "id": c.id,
            "user": c.user.user_id or c.user.phone,
            "user_name": f"{c.user.first_name} {c.user.last_name}".strip() or c.user.phone,
            "image": request.build_absolute_uri(c.image.url) if c.image else None,
            "weight": c.weight, "length": c.length, "width": c.width, "height": c.height,
            "comment": c.comment or "",
            "price": float(c.price) if c.price else None,
            "admin_note": c.admin_note or "",
            "is_responded": c.is_responded,
            "created_at": c.created_at.strftime("%d.%m.%Y %H:%M"),
        })
    return Response({"results": data})


@api_view(["PATCH"])
@permission_classes([IsAdminUser])
def calc_reply(request, pk):
    try:
        c = CalculationRequest.objects.get(pk=pk)
    except CalculationRequest.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)
    c.price = request.data.get("price")
    c.admin_note = request.data.get("admin_note", "")
    c.is_responded = True
    c.save()
    return Response({"ok": True})


# ── KODSIZ TOVARLAR ──────────────────────────

@api_view(["POST"])
@permission_classes([IsAdminUser])
@parser_classes([MultiPartParser, FormParser])
def unassigned_import_excel(request):
    """POST /api/admin/unassigned/import/ — Excel import (TREK RAQAM, REYS, SANA, IZOH)"""
    import openpyxl
    from io import BytesIO
    from django.utils.dateparse import parse_date

    file = request.FILES.get("file")
    if not file:
        return Response({"error": "Fayl yuklanmagan"}, status=400)

    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return Response({"error": f"Faylni o'qib bo'lmadi: {e}"}, status=400)

    # Ustun sarlavhalarini topish
    headers = {}
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_vals = [str(v).strip().upper() if v else "" for v in row]
        for j, val in enumerate(row_vals):
            if "TREK" in val: headers["track_code"] = j
            if "REYS" in val: headers["flight_name"] = j
            if "SANA" in val: headers["created_at"] = j
            if "IZOH" in val: headers["note"] = j
        if "track_code" in headers:
            header_row = i
            break

    if "track_code" not in headers:
        return Response({"error": "TREK RAQAM ustuni topilmadi"}, status=400)

    imported = 0
    skipped = 0
    errors = []

    for row in ws.iter_rows(min_row=(header_row or 1) + 1, values_only=True):
        track = row[headers["track_code"]] if len(row) > headers["track_code"] else None
        if not track or str(track).strip().lower() in ["", "none"]:
            continue

        track = str(track).strip()
        flight = str(row[headers["flight_name"]]).strip() if headers.get("flight_name") is not None and len(row) > headers["flight_name"] and row[headers["flight_name"]] else ""
        note = str(row[headers["note"]]).strip() if headers.get("note") is not None and len(row) > headers["note"] and row[headers["note"]] else ""

        created_at = timezone.now()
        if headers.get("created_at") is not None and len(row) > headers["created_at"]:
            val = row[headers["created_at"]]
            if val:
                try:
                    if hasattr(val, "date"):
                        created_at = timezone.make_aware(
                            timezone.datetime.combine(val.date() if hasattr(val, "date") else val, timezone.datetime.min.time())
                        )
                except Exception:
                    pass

        obj, created = UnassignedCargo.objects.get_or_create(
            track_code=track,
            defaults={"flight_name": flight, "note": note, "created_at": created_at}
        )
        if created:
            imported += 1
        else:
            # Yangilaymiz
            if flight: obj.flight_name = flight
            if note: obj.note = note
            obj.save()
            skipped += 1

    return Response({"ok": True, "imported": imported, "updated": skipped, "errors": errors})


@api_view(["GET"])
@permission_classes([IsAdminUser])
def unassigned_list(request):
    qs = UnassignedCargo.objects.order_by("-created_at")
    if q := request.GET.get("search", "").strip():
        qs = qs.filter(Q(track_code__icontains=q) | Q(flight_name__icontains=q))
    total = qs.count()
    page = int(request.GET.get("page", 1))
    size = 30
    results = [{"id": c.id, "track_code": c.track_code, "flight_name": c.flight_name, "note": c.note or "", "created_at": c.created_at.strftime("%d.%m.%Y")} for c in qs[(page - 1) * size: page * size]]
    return Response({"results": results, "total": total, "page": page, "pages": (total + size - 1) // size})


@api_view(["POST"])
@permission_classes([IsAdminUser])
def unassigned_create(request):
    try:
        c = UnassignedCargo.objects.create(
            track_code=request.data["track_code"],
            flight_name=request.data["flight_name"],
            note=request.data.get("note", ""),
        )
        return Response({"ok": True, "id": c.id})
    except Exception as e:
        return Response({"error": str(e)}, status=400)


@api_view(["DELETE"])
@permission_classes([IsAdminUser])
def unassigned_delete(request, pk):
    try:
        UnassignedCargo.objects.get(pk=pk).delete()
        return Response({"ok": True})
    except UnassignedCargo.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)


# ── BILDIRISHNOMALAR ─────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def notification_list(request):
    qs = Notification.objects.select_related("user").order_by("-created_at")
    total = qs.count()
    page = int(request.GET.get("page", 1))
    size = 30
    results = []
    for n in qs[(page - 1) * size: page * size]:
        results.append({
            "id": n.id,
            "user": n.user.user_id or n.user.phone,
            "user_name": f"{n.user.first_name} {n.user.last_name}".strip() or n.user.phone,
            "title": n.title, "body": n.body,
            "notification_type": n.notification_type,
            "is_read": n.is_read,
            "created_at": n.created_at.strftime("%d.%m.%Y %H:%M"),
        })
    return Response({"results": results, "total": total, "page": page, "pages": (total + size - 1) // size})


@api_view(["POST"])
@permission_classes([IsAdminUser])
def notification_send(request):
    """Bildirishnoma yuborish — hammaga yoki bitta foydalanuvchiga"""
    title = request.data.get("title", "")
    body = request.data.get("body", "")
    user_id_str = request.data.get("user_id", "").strip()
    notif_type = request.data.get("notification_type", "Umumiy")

    if user_id_str:
        users = User.objects.filter(user_id=user_id_str, is_staff=False)
    else:
        users = User.objects.filter(is_staff=False, status="approved")

    count = 0
    for u in users:
        Notification.objects.create(user=u, title=title, body=body, notification_type=notif_type)
        count += 1

    # Firebase push (agar fcm_token bo'lsa)
    try:
        from utils.push_service import send_push_notification
        for u in users:
            if u.fcm_token:
                send_push_notification(u.fcm_token, title, body)
    except Exception:
        pass

    return Response({"ok": True, "sent_to": count})


# ── VIDEO DARSLIK ────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminUser])
def video_list(request):
    videos = TutorialVideo.objects.order_by("-created_at")
    return Response([{"id": v.id, "video_url": v.video_url, "created_at": v.created_at.strftime("%d.%m.%Y")} for v in videos])


@api_view(["POST"])
@permission_classes([IsAdminUser])
def video_create(request):
    url = request.data.get("video_url", "")
    if not url:
        return Response({"error": "URL kiritilmagan"}, status=400)
    v = TutorialVideo.objects.create(video_url=url)
    return Response({"ok": True, "id": v.id})


@api_view(["DELETE"])
@permission_classes([IsAdminUser])
def video_delete(request, pk):
    try:
        TutorialVideo.objects.get(pk=pk).delete()
        return Response({"ok": True})
    except TutorialVideo.DoesNotExist:
        return Response({"error": "Topilmadi"}, status=404)


# ── SOZLAMALAR ───────────────────────────────

@api_view(["GET", "PATCH"])
@permission_classes([IsAdminUser])
def settings_view(request):
    obj, _ = WarehouseSettings.objects.get_or_create(pk=1)
    if request.method == "GET":
        return Response({
            "china_avia_phone": obj.china_avia_phone, "china_avia_address": obj.china_avia_address,
            "china_avia_price": float(obj.china_avia_price), "china_avia_term": obj.china_avia_term,
            "china_auto_phone": obj.china_auto_phone, "china_auto_address": obj.china_auto_address,
            "china_auto_price": float(obj.china_auto_price), "china_auto_term": obj.china_auto_term,
            "contact_telegram": obj.contact_telegram, "contact_instagram": obj.contact_instagram,
            "contact_phone": obj.contact_phone,
        })
    for f in request.data:
        if hasattr(obj, f): setattr(obj, f, request.data[f])
    obj.save()
    return Response({"ok": True})


@api_view(["GET", "PATCH"])
@permission_classes([IsAdminUser])
def app_version_view(request):
    obj, _ = AppVersion.objects.get_or_create(pk=1, defaults={"version": "1.0.0", "play_store_url": "", "app_store_url": ""})
    if request.method == "GET":
        return Response({"version": obj.version, "play_store_url": obj.play_store_url, "app_store_url": obj.app_store_url, "is_force_update": obj.is_force_update})
    for f in ["version", "play_store_url", "app_store_url", "is_force_update"]:
        if f in request.data: setattr(obj, f, request.data[f])
    obj.save()
    return Response({"ok": True})
